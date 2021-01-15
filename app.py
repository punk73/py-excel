# from openpyxl import Workbook
# wb = Workbook()

# # grab the active worksheet
# ws = wb.active

# # Data can be assigned directly to cells
# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# # Save the file
# wb.save("sample.xlsx")
import openpyxl
import csv

file = open("SF92D_R_PLN04_20200828.CSV", "r")

csv_file = csv.DictReader(file)
list_csv = list(csv_file)

print(list_csv[0])